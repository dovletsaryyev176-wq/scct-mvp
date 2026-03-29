from flask import jsonify, request, session, send_file
import io
import openpyxl
from datetime import date, datetime
from decimal import Decimal
from db import Db
from decorators import roles_required
from all_types_description import OrderStatuses, PaymentTypes
from . import accounter_bp

# -------------------------------------------------------------
# 1. Список курьеров с долгами (ожидающих кассы на дату)
# -------------------------------------------------------------
@accounter_bp.route('/couriers/debt', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def get_couriers_debt():
    target_date_str = request.args.get('date')
    courier_name_filter = request.args.get('courier_name')
    
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты. Ожидается YYYY-MM-DD'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            # 1. Получаем список всех курьеров, у которых были заказы ИЛИ есть записи о платежах в эту дату
            # Проще собрать статистику по таблице заказов + таблице платежей, сгруппировав по курьеру.
            
            # Узнаем общее число заказов и число исполненных заказов курьера за день
            orders_sql = """
                SELECT 
                    o.courier_id, 
                    u.full_name as courier_name,
                    COUNT(o.id) as total_orders,
                    SUM(CASE WHEN o.status = %s THEN 1 ELSE 0 END) as completed_orders
                FROM orders o
                JOIN users u ON o.courier_id = u.id
                WHERE o.delivery_date = %s
                GROUP BY o.courier_id, u.full_name
            """
            cursor.execute(orders_sql, (OrderStatuses.DELIVERED, target_date))
            orders_stats = cursor.fetchall()

            # Узнаем долги (только то, что физически у них на руках: is_handed_over = FALSE)
            # При этом группируем по типу оплаты (cash / cash_and_card / card - хотя card по сути безнал онлайн)
            debt_sql = """
                SELECT 
                    cp.courier_id,
                    cp.payment_type,
                    SUM(cp.cash_amount) as total_cash,
                    SUM(cp.card_amount) as total_card
                FROM courier_payments cp
                JOIN orders o ON cp.order_id = o.id
                WHERE o.delivery_date = %s AND cp.is_handed_over = FALSE
                GROUP BY cp.courier_id, cp.payment_type
            """
            cursor.execute(debt_sql, (target_date,))
            debt_stats = cursor.fetchall()

            # Собираем все в один ответ
            couriers_dict = {}
            for stat in orders_stats:
                cid = stat['courier_id']
                couriers_dict[cid] = {
                    'courier_id': cid,
                    'courier_name': stat['courier_name'],
                    'total_orders': int(stat['total_orders']),
                    'completed_orders': int(stat['completed_orders']),
                    'debts_by_type': {},
                    'total_cash_debt': 0.0,
                    'total_card_debt': 0.0,
                    'total_debt': 0.0
                }

            # Могут быть курьеры, у которых нет новых заказов на эту дату,
            # но есть платежи (если они выполнили вчерашний заказ сегодня).
            for d in debt_stats:
                cid = d['courier_id']
                if cid not in couriers_dict:
                    # Узнаем имя курьера
                    cursor.execute("SELECT full_name FROM users WHERE id = %s", (cid,))
                    u_row = cursor.fetchone()
                    couriers_dict[cid] = {
                        'courier_id': cid,
                        'courier_name': u_row['full_name'] if u_row else f"Курьер ID {cid}",
                        'total_orders': 0,
                        'completed_orders': 0,
                        'debts_by_type': {},
                        'total_cash_debt': 0.0,
                        'total_card_debt': 0.0,
                        'total_debt': 0.0
                    }

                p_type = d['payment_type']
                cash_sum = float(d['total_cash']) if d['total_cash'] else 0.0
                card_sum = float(d['total_card']) if d['total_card'] else 0.0

                couriers_dict[cid]['debts_by_type'][p_type] = {
                    'cash': cash_sum,
                    'card': card_sum
                }

                couriers_dict[cid]['total_cash_debt'] += cash_sum
                couriers_dict[cid]['total_card_debt'] += card_sum
                couriers_dict[cid]['total_debt'] += (cash_sum + card_sum)

            result_list = list(couriers_dict.values())
            
            if courier_name_filter:
                search_term = courier_name_filter.lower()
                result_list = [c for c in result_list if search_term in c['courier_name'].lower()]

            # Сортируем по имени курьера
            result_list.sort(key=lambda x: x['courier_name'])

            return jsonify({
                'date': target_date.isoformat(),
                'couriers': result_list
            }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 2. Детализация платежей и заказов конкретного курьера за день
# -------------------------------------------------------------
@accounter_bp.route('/couriers/<int:courier_id>/payments', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def get_courier_payments_details(courier_id):
    target_date_str = request.args.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            # Получаем детальный список
            details_sql = """
                SELECT 
                    cp.id as payment_id,
                    cp.payment_type as payment_collected_type,
                    cp.cash_amount,
                    cp.card_amount,
                    cp.created_at as payment_date,
                    cp.is_handed_over,
                    cp.handed_over_at,
                    cp.accounter_note,
                    
                    o.id as order_id,
                    o.client_id,
                    c.full_name as client_name,
                    cphone.phone as client_phone,
                    ca.address_line as client_address,
                    s.name as street_name,
                    ca.appartment,
                    ca.entrance,
                    ca.floor,
                    o.status as order_status,
                    o.total_amount as order_total_amount
                FROM courier_payments cp
                JOIN orders o ON cp.order_id = o.id
                LEFT JOIN clients c ON o.client_id = c.id
                LEFT JOIN client_phones cphone ON o.client_phone_id = cphone.id
                LEFT JOIN client_addresses ca ON o.client_address_id = ca.id
                LEFT JOIN streets s ON ca.street_id = s.id
                WHERE cp.courier_id = %s AND o.delivery_date = %s
                ORDER BY cp.created_at DESC
            """
            cursor.execute(details_sql, (courier_id, target_date))
            details = cursor.fetchall()

            for d in details:
                d['cash_amount'] = float(d['cash_amount'])
                d['card_amount'] = float(d['card_amount'])
                d['order_total_amount'] = float(d['order_total_amount']) if d.get('order_total_amount') is not None else 0.0
                if d['payment_date']:
                    d['payment_date'] = d['payment_date'].isoformat()
                if d['handed_over_at']:
                    d['handed_over_at'] = d['handed_over_at'].isoformat()

            return jsonify({
                'date': target_date.isoformat(),
                'courier_id': courier_id,
                'details': details
            }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 3. Принять кассу (Передача денег бухгалтеру)
# -------------------------------------------------------------
@accounter_bp.route('/couriers/<int:courier_id>/handover', methods=['POST'])
@roles_required('admin', 'accounter')
def accept_courier_handover(courier_id):
    accounter_id = session.get('user_id')
    
    data = request.get_json() or {}
    target_date_str = data.get('date') # Если хотят закрыть долг за вчерашний день
    note = data.get('note') # Заметка бухгалтера

    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        conn.begin()
        with conn.cursor() as cursor:
            # Находим записи, которые еще не сданы
            find_sql = """
                SELECT cp.id 
                FROM courier_payments cp
                JOIN orders o ON cp.order_id = o.id
                WHERE cp.courier_id = %s AND o.delivery_date = %s AND cp.is_handed_over = FALSE
                FOR UPDATE
            """
            cursor.execute(find_sql, (courier_id, target_date))
            pending_records = cursor.fetchall()

            if not pending_records:
                conn.rollback()
                return jsonify({'error': 'У данного курьера нет несданных платежей на указанную дату'}), 400

            ids = [r['id'] for r in pending_records]
            placeholders = ', '.join(['%s'] * len(ids))
            
            # Обновляем статусы
            update_sql = f"""
                UPDATE courier_payments 
                SET is_handed_over = TRUE, 
                    handed_over_at = NOW(), 
                    accounter_id = %s,
                    accounter_note = %s
                WHERE id IN ({placeholders})
            """
            params = [accounter_id, note] + ids
            cursor.execute(update_sql, tuple(params))
            
            conn.commit()
            
            return jsonify({
                'message': 'Деньги успешно приняты',
                'courier_id': courier_id,
                'date': target_date.isoformat(),
                'records_updated': len(ids)
            }), 200

    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 4. Итоговая сумма к сдаче по всем курьерам за день
# -------------------------------------------------------------
@accounter_bp.route('/debt/summary', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def get_all_couriers_debt_summary():
    target_date_str = request.args.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                SELECT 
                    SUM(cp.cash_amount) as total_cash,
                    SUM(cp.card_amount) as total_card
                FROM courier_payments cp
                JOIN orders o ON cp.order_id = o.id
                WHERE o.delivery_date = %s AND cp.is_handed_over = FALSE
            """
            cursor.execute(sql, (target_date,))
            res = cursor.fetchone()

            cash = float(res['total_cash']) if res and res['total_cash'] is not None else 0.0
            card = float(res['total_card']) if res and res['total_card'] is not None else 0.0

            return jsonify({
                'date': target_date.isoformat(),
                'total_cash_debt': cash,
                'total_card_debt': card,
                'total_debt': cash + card
            }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 5. Экспорт списка долгов курьеров в Excel
# -------------------------------------------------------------
@accounter_bp.route('/couriers/debt/export', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def export_couriers_debt_excel():
    target_date_str = request.args.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты. Ожидается YYYY-MM-DD'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            # Тот же запрос на статистику, что и в /couriers/debt
            orders_sql = """
                SELECT 
                    o.courier_id, 
                    u.full_name as courier_name,
                    COUNT(o.id) as total_orders,
                    SUM(CASE WHEN o.status = %s THEN 1 ELSE 0 END) as completed_orders
                FROM orders o
                JOIN users u ON o.courier_id = u.id
                WHERE o.delivery_date = %s
                GROUP BY o.courier_id, u.full_name
            """
            cursor.execute(orders_sql, (OrderStatuses.DELIVERED, target_date))
            orders_stats = cursor.fetchall()

            debt_sql = """
                SELECT 
                    cp.courier_id,
                    cp.payment_type,
                    SUM(cp.cash_amount) as total_cash,
                    SUM(cp.card_amount) as total_card
                FROM courier_payments cp
                JOIN orders o ON cp.order_id = o.id
                WHERE o.delivery_date = %s AND cp.is_handed_over = FALSE
                GROUP BY cp.courier_id, cp.payment_type
            """
            cursor.execute(debt_sql, (target_date,))
            debt_stats = cursor.fetchall()

            couriers_dict = {}
            for stat in orders_stats:
                cid = stat['courier_id']
                couriers_dict[cid] = {
                    'courier_name': stat['courier_name'],
                    'total_orders': int(stat['total_orders']),
                    'completed_orders': int(stat['completed_orders']),
                    'total_cash_debt': 0.0,
                    'total_card_debt': 0.0,
                    'total_debt': 0.0
                }

            for d in debt_stats:
                cid = d['courier_id']
                if cid not in couriers_dict:
                    cursor.execute("SELECT full_name FROM users WHERE id = %s", (cid,))
                    u_row = cursor.fetchone()
                    couriers_dict[cid] = {
                        'courier_name': u_row['full_name'] if u_row else f"Курьер ID {cid}",
                        'total_orders': 0,
                        'completed_orders': 0,
                        'total_cash_debt': 0.0,
                        'total_card_debt': 0.0,
                        'total_debt': 0.0
                    }

                cash_sum = float(d['total_cash']) if d['total_cash'] else 0.0
                card_sum = float(d['total_card']) if d['total_card'] else 0.0

                couriers_dict[cid]['total_cash_debt'] += cash_sum
                couriers_dict[cid]['total_card_debt'] += card_sum
                couriers_dict[cid]['total_debt'] += (cash_sum + card_sum)

            result_list = list(couriers_dict.values())
            result_list.sort(key=lambda x: x['courier_name'])

            # Формирование Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Долги курьеров"

            headers = ["Курьер", "Всего заказов", "Завершено", "К сдаче (Наличные)", "К сдаче (Карта)", "Итого к сдаче"]
            ws.append(headers)

            # Настраиваем ширину колонок для красоты
            column_widths = [30, 15, 15, 20, 20, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            for c in result_list:
                ws.append([
                    c['courier_name'],
                    c['total_orders'],
                    c['completed_orders'],
                    c['total_cash_debt'],
                    c['total_card_debt'],
                    c['total_debt']
                ])

            # Сохраняем в память
            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            filename = f"couriers_debt_{target_date.strftime('%Y-%m-%d')}.xlsx"

            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 6. Отчет по выданным продуктам (услугам)
# -------------------------------------------------------------
@accounter_bp.route('/movements/summary', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def get_movements_summary():
    target_date_str = request.args.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты. Ожидается YYYY-MM-DD'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            # Учитываем скидку: сумма по позициям = доля позиции в заказе * итог заказа (o.total_amount)
            sql = """
                SELECT 
                    s.name as service_name,
                    pt.name as price_type_name,
                    o.payment_type,
                    oi.quantity,
                    oi.total_price,
                    o.total_amount as order_total,
                    SUM(oi.total_price) OVER (PARTITION BY o.id) as order_items_sum
                FROM orders o
                JOIN order_items oi ON o.id = oi.order_id
                JOIN services s ON oi.service_id = s.id
                JOIN clients c ON o.client_id = c.id
                JOIN price_types pt ON c.price_type_id = pt.id
                WHERE o.delivery_date = %s AND o.status = %s
            """
            cursor.execute(sql, (target_date, OrderStatuses.DELIVERED))
            rows = cursor.fetchall()

            grouped_data = {}
            for row in rows:
                key = (row['service_name'], row['price_type_name'])
                if key not in grouped_data:
                    grouped_data[key] = {
                        'service_name': row['service_name'],
                        'price_type_name': row['price_type_name'],
                        'total_quantity': 0.0,
                        'amounts_by_payment': {ptype: 0.0 for ptype in PaymentTypes.CHOICES},
                        'total_amount': 0.0
                    }

                order_total = float(row['order_total']) if row['order_total'] else 0.0
                items_sum = float(row['order_items_sum']) if row['order_items_sum'] else 0.0
                item_price = float(row['total_price']) if row['total_price'] else 0.0
                ratio = (order_total / items_sum) if items_sum else 0.0
                amt = item_price * ratio
                qty = float(row['quantity']) if row['quantity'] else 0.0

                grouped_data[key]['total_quantity'] += qty
                grouped_data[key]['total_amount'] += amt
                if row['payment_type'] in grouped_data[key]['amounts_by_payment']:
                    grouped_data[key]['amounts_by_payment'][row['payment_type']] += amt

            result_list = list(grouped_data.values())
            # Сортировка по услуге и типу цены
            result_list.sort(key=lambda x: (x['service_name'], x['price_type_name']))

            return jsonify({
                'date': target_date.isoformat(),
                'movements': result_list
            }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

# -------------------------------------------------------------
# 7. Экспорт отчета по выданным продуктам в Excel
# -------------------------------------------------------------
@accounter_bp.route('/movements/summary/export', methods=['GET'])
@roles_required('admin', 'operator', 'accounter')
def export_movements_summary_excel():
    target_date_str = request.args.get('date')
    if target_date_str:
        try:
            target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Неправильный формат даты. Ожидается YYYY-MM-DD'}), 400
    else:
        target_date = date.today()

    conn = Db.get_connection()
    try:
        with conn.cursor() as cursor:
            sql = """
                SELECT 
                    s.name as service_name,
                    pt.name as price_type_name,
                    o.payment_type,
                    oi.quantity,
                    oi.total_price,
                    o.total_amount as order_total,
                    SUM(oi.total_price) OVER (PARTITION BY o.id) as order_items_sum
                FROM orders o
                JOIN order_items oi ON o.id = oi.order_id
                JOIN services s ON oi.service_id = s.id
                JOIN clients c ON o.client_id = c.id
                JOIN price_types pt ON c.price_type_id = pt.id
                WHERE o.delivery_date = %s AND o.status = %s
            """
            cursor.execute(sql, (target_date, OrderStatuses.DELIVERED))
            rows = cursor.fetchall()

            grouped_data = {}
            for row in rows:
                key = (row['service_name'], row['price_type_name'])
                if key not in grouped_data:
                    grouped_data[key] = {
                        'service_name': row['service_name'],
                        'price_type_name': row['price_type_name'],
                        'total_quantity': 0.0,
                        'amounts_by_payment': {ptype: 0.0 for ptype in PaymentTypes.CHOICES},
                        'total_amount': 0.0
                    }

                order_total = float(row['order_total']) if row['order_total'] else 0.0
                items_sum = float(row['order_items_sum']) if row['order_items_sum'] else 0.0
                item_price = float(row['total_price']) if row['total_price'] else 0.0
                ratio = (order_total / items_sum) if items_sum else 0.0
                amt = item_price * ratio
                qty = float(row['quantity']) if row['quantity'] else 0.0

                grouped_data[key]['total_quantity'] += qty
                grouped_data[key]['total_amount'] += amt
                if row['payment_type'] in grouped_data[key]['amounts_by_payment']:
                    grouped_data[key]['amounts_by_payment'][row['payment_type']] += amt

            result_list = list(grouped_data.values())
            result_list.sort(key=lambda x: (x['service_name'], x['price_type_name']))

            # format=json — вернуть те же данные JSON для проверки в Insomnia/Postman
            if request.args.get('format') == 'json':
                return jsonify({
                    'date': target_date.isoformat(),
                    'movements': result_list
                }), 200

            # Формирование Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Выданные продукты"

            payment_labels = [PaymentTypes.LABELS[pt]['ru'] for pt in PaymentTypes.CHOICES]
            headers = ["Вид услуги", "Тип цены", "Общее кол-во"] + [f"Сумма ({lbl})" for lbl in payment_labels] + ["Итого Сумма"]
            ws.append(headers)

            column_widths = [30, 20, 15] + [20] * len(PaymentTypes.CHOICES) + [20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            for r in result_list:
                row_data = [
                    r['service_name'],
                    r['price_type_name'],
                    r['total_quantity']
                ]
                for pt in PaymentTypes.CHOICES:
                    row_data.append(r['amounts_by_payment'][pt])
                row_data.append(r['total_amount'])
                
                ws.append(row_data)

            excel_file = io.BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)

            filename = f"movements_summary_{target_date.strftime('%Y-%m-%d')}.xlsx"

            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()
